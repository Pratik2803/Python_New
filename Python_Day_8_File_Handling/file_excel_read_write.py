from openpyxl import Workbook, load_workbook

# Create a workbook and add multiple sheets
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = "Pratik"

# Write data in sheet 1
sheet1["A1"] = "Name"
sheet1["B1"] = "Age"
sheet1["C1"] = "City"

sheet1.append(['Pratik', 37, 'Noida'])
sheet1.append(['Pratik', 37, 'Greater Noida West'])

sheet2 = workbook.create_sheet(title="Kiara", index=1)

# Write data in sheet 2
sheet2["A1"] = "Name"
sheet2["B1"] = "Age"
sheet2["C1"] = "City"

sheet2.append(['Kiara', 87, 'Noida'])
sheet2.append(['Kiara', 87, 'Greater Noida West'])


workbook.save("Timepass.xlsx")

workbook = load_workbook(filename="Timepass.xlsx")

sheets = workbook.sheetnames

for sheet in sheets:
    active_sheet = workbook[sheet]
    for row in active_sheet.iter_rows(values_only=True):
        print(row)
